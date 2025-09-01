import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class AutomacaoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Automação de Formulários/Tabelas")
        self.root.geometry("600x450")

        # --- Links
        tk.Label(root, text="Link do Formulário:").pack(anchor="w", padx=10, pady=5)
        self.link_entry = tk.Entry(root, width=70)
        self.link_entry.pack(padx=10, pady=5)
        
        # -- Arquivo
        tk.Button(root, text="Selecionar Arquivo Excel", command=self.selecionar_arquivo).pack(pady=5)
        tk.Button(root, text="Limpar arquivo", command=self.limpar_arquivo).pack(pady=5)
        self.arquivo_label = tk.Label(root, text="Nenhum arquivo selecionado", fg="gray")
        self.arquivo_label.pack()

        # -- Dados
        tk.Label(root, text="Dados para preenchimento (formato: Nome,Categoria,Tamanho,Preço,Quantidade)").pack(anchor="w", padx=10, pady=5)
        self.dados_text = tk.Text(root, height=8, width=70)
        self.dados_text.pack(padx=10, pady=5)

        # -- Botão executar
        tk.Button(root, text="Executar Automação", command=self.executar).pack(pady=10)

        # -- Status da automação
        self.status_label = tk.Label(root, text="Aguardando ação...", fg="blue")
        self.status_label.pack(pady=5)

        self.caminho_arquivo = None

    def selecionar_arquivo(self):
        arquivo = filedialog.askopenfilename(
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if arquivo:
            self.caminho_arquivo = arquivo
            self.arquivo_label.config(text=f"Selecionado: {arquivo}")

    def executar(self):
        link = self.link_entry.get().strip()
        dados = self.dados_text.get("1.0", tk.END).strip()

        if not link and not self.caminho_arquivo:
            messagebox.showwarning("Aviso", "Informe um link ou selecione um arquivo.")
            return
        
        if not dados:
            messagebox.showwarning("Aviso", "Adicione os dados para preenchimento.")
            return
        
        self.status_label.config(text="Executando automação...", fg="green")
        try:
            if link:  
                self.preencher_formulario(link, dados)
            elif self.caminho_arquivo:  
                arquivo_final = self.preencher_excel(self.caminho_arquivo, dados)
                self.status_label.config(text=f"Arquivo salvo em {arquivo_final}", fg="darkgreen")

        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro: {e}")
            self.status_label.config(text="Erro na automação!", fg="red")
        else:
            self.status_label.config(text="Automação concluída!", fg="darkgreen")

    def limpar_arquivo(self):
        self.caminho_arquivo = None
        self.arquivo_label.config(text="Nenhum arquivo selecionado", fg="gray")

    def preencher_formulario(self, link, dados):
        driver = webdriver.Chrome()
        try:
            driver.get(link)

    
            messagebox.showinfo(
                "Login necessário",
                "Por favor, faça o login e clique em OK quando estiver pronto."
            )

            linhas = [l.strip() for l in dados.split("\n") if l.strip()]  
            for linha in linhas:
                campos_item = linha.split(",")
                if len(campos_item) != 5:
                    messagebox.showwarning("Aviso", f"Linha inválida: {linha}\nEsperado: Nome,Categoria,Tamanho,Preço,Quantidade")
                    continue

                nome, categoria, tamanho, preco, quantidade = campos_item

               
                driver.find_element(By.XPATH, "//input[@name='nome']").send_keys(nome)
                driver.find_element(By.XPATH, "//input[@name='categoria']").send_keys(categoria)

                if categoria.lower() in ["camisa", "calça", "short", "jaqueta"]:
                    driver.find_element(By.XPATH, "//input[@name='tamanho']").send_keys(tamanho)

                driver.find_element(By.XPATH, "//input[@name='preco']").send_keys(preco)
                driver.find_element(By.XPATH, "//input[@name='quantidade']").send_keys(quantidade)

                driver.find_element(By.XPATH, "//button[text()='Salvar']").click()

              
                try:
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//button[text()='OK']"))
                    ).click()
                except:
                    print("Não encontrou a mensagem de OK, continuando...")

                time.sleep(0.5)  

        except Exception as e:
            messagebox.showerror("Erro", f"Erro na automação do formulário: {e}")

    def preencher_excel(self, caminho_arquivo, dados):
        wb = openpyxl.load_workbook(caminho_arquivo)
        ws = wb.active

        cabecalho = ["Nome", "Categoria", "Tamanho", "Preço", "Quantidade"]
        for col, titulo in enumerate(cabecalho, start=1):
            cell = ws.cell(row=1, column=col, value=titulo)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        linhas = [l.strip() for l in dados.split("\n") if l.strip()]
        for i, linha in enumerate(linhas, start=2):
            campos = linha.split(",")
            if len(campos) != 5:
                continue 
            for j, valor in enumerate(campos, start=1):
                ws.cell(row=i, column=j, value=valor)

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 5

        novo_arquivo = caminho_arquivo.replace(".xlsx", "_preenchido.xlsx")
        wb.save(novo_arquivo)
        return novo_arquivo


if __name__ == "__main__":
    root = tk.Tk()
    app = AutomacaoApp(root)
    root.mainloop()